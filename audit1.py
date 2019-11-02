import os
import openpyxl
import sys
import sftp

print("Start of Program")

if len(sys.argv) != 2:
  exit("Please enter an argument")

input_filename = sys.argv[1]


HOST = "54.80.104.5"
USER = "student"
KEY = "dnp423_student.pem"

if not os.path.isfile(KEY):
  exit("RSA key not found")
else:
  print("RSA key found!")

SUCCESS, SSHCLIENT = sftp.setup_ssh_connection(HOST, USER, KEY)

print(SUCCESS, SSHCLIENT)

if SUCCESS:
  SUCCESS, FTPCLIENT = sftp.setup_sftp_sesstion(SSHCLIENT)

  if SUCCESS:
    SUCCESS = sftp.get_file(FTPCLIENT, input_filename)


sftp.close_ssh_connection(SSHCLIENT)

print("End of Program")



working_folder = os.getcwd()
# print(working_folder)

contents = os.listdir(working_folder)
# print(contents)

# for i in contents:
#    print(i)

WB = openpyxl.Workbook()

WS1 = WB.active
ws1_name = input_filename.replace(".txt", "")
WS1.title = ws1_name

WS2 = WB.create_sheet("WORD_ANALYSIS")
WS3 = WB.create_sheet("SUMMARY")

wb_filename = input_filename.replace(".txt", ".xlsx")

WS1.cell(1,1).value = "Line_NUM"
WS1.cell(1,2).value = "WORD_COUNT"
WS1.cell(1,3).value = "CHAR_COUNT"
WS1.cell(1,4).value = "NOWS_CHAR_COUNT"
WS1.cell(1,5).value = "LINE_CONTENT"

WS2.cell(1,1).value = "Number of Lines"
WS2.cell(1,2).value = "Total Word Count"
WS2.cell(1,3).value = "Total Char Count"
WS2.cell(1,4).value = "Total No Whitespace Char Count"

WS3.cell(1,1).value = "Word Count"
WS3.cell(1,2).value = "Word Length"

WORD_TABLE = []
NUM_COUNTERS = 17
ws3_row = 2

print(len(WORD_TABLE), WORD_TABLE)

for my_count in range(0, NUM_COUNTERS, 1):
  WORD_TABLE.append(0)
  WS3.cell(ws3_row,2).value = my_count
  ws3_row += 1
print(len(WORD_TABLE), WORD_TABLE)


file_object = open(input_filename, mode='r')
# print(file_object)

line_num = 0
total_word_count =0
total_char_count =0
total_nows_char_count =0
for i in file_object:

  line_num+=1

  line_char_count = len(i)
  total_char_count += line_char_count

  my_word_list = i.split()
  line_word_count = len(my_word_list)
  total_word_count += line_word_count

  line_nows_char_count = 0
  for j in my_word_list:
    line_nows_char_count += len(j)
    if len(j) >= NUM_COUNTERS -1:
      WORD_TABLE[NUM_COUNTERS-1] +=1

      WS3.cell(NUM_COUNTERS-1,1).value = WORD_TABLE[NUM_COUNTERS-1]

    else:
      WORD_TABLE[len(j)] += 1
      WS3.cell(len(j) +2 ,1).value = WORD_TABLE[len(j)]


  # print(my_word_list)

  total_nows_char_count += line_nows_char_count
  WS1.cell(line_num+1,1).value = line_num
  WS1.cell(line_num+1,2).value = line_word_count
  WS1.cell(line_num+1,3).value = line_char_count
  WS1.cell(line_num+1,4).value = line_nows_char_count
  WS1.cell(line_num+1,5).value = i

print(len(WORD_TABLE), WORD_TABLE)

data = openpyxl.chart.Reference(WS3, min_col=1, min_row=1, max_row=NUM_COUNTERS+1)
category = openpyxl.chart.Reference(WS3, min_col=2, min_row=2, max_row=NUM_COUNTERS+1)

pie = openpyxl.chart.PieChart()

pie.add_data(data, titles_from_data=True)
pie.set_categories(category)

pie.title = "WORD COUNT SUMMARY"
WS3.add_chart(pie,"F1")

WS2.cell(2,1).value = line_num
WS2.cell(2,2).value = total_word_count
WS2.cell(2,3).value = total_char_count
WS2.cell(2,4).value = total_nows_char_count
  
WB.save(wb_filename)

#   print(line_num, line_word_count, line_char_count, line_nows_char_count, i, end="")
# print()
# print("total word count ---", total_word_count)
# print("total char count ---", total_char_count)
# print("total no ws char count ---", total_nows_char_count)
